# app/report_generator.py
from __future__ import annotations

import io
import json
import sqlite3
from pathlib import Path

import openpyxl
from openpyxl.drawing.image import Image as OpenPyxlImage
from openpyxl.styles import Alignment, Font, PatternFill
from PIL import Image, ImageDraw

from app.config import PHOTOS_DIR, REPORTS_DIR, get_local_ip, load_config
from app.database import get_connection
from app.repositories.settings_repository import SettingsRepository


def _draw_face_circle(image_path: Path, bbox_json: str) -> io.BytesIO | None:
    """Safely extracts a face crop from the original image and draws a target circle."""
    try:
        if not image_path.exists():
            return None

        bbox = json.loads(bbox_json)  # [x1, y1, x2, y2]
        img = Image.open(image_path).convert("RGB")

        # Crop to the face with a slight margin
        margin = 40
        left = max(0, bbox[0] - margin)
        top = max(0, bbox[1] - margin)
        right = min(img.width, bbox[2] + margin)
        bottom = min(img.height, bbox[3] + margin)

        face_img = img.crop((left, top, right, bottom))

        # Draw target ellipse
        draw = ImageDraw.Draw(face_img)
        draw.ellipse(
            [(10, 10), (face_img.width - 10, face_img.height - 10)],
            outline="green",
            width=6,
        )

        # Resize for Excel embedding
        face_img.thumbnail((80, 80))
        out = io.BytesIO()
        face_img.save(out, format="PNG")
        out.seek(0)
        return out
    except Exception:
        return None


def _find_face_evidence_for_session(db: sqlite3.Connection, worker_id: int, date: str, session: str):
    """Returns evidence strictly for the given session — no cross-session
    fallback. This is what lets Morning and Evening evidence be shown
    separately instead of one photo standing in for both."""
    return db.execute(
        """
        SELECT f.bbox, p.filename, p.session, p.date
        FROM attendance_faces f
        JOIN photos_log p ON f.photo_log_id = p.id
        WHERE f.worker_id = ? AND f.date = ? AND f.session = ?
        ORDER BY f.created_at DESC LIMIT 1
        """,
        (worker_id, date, session),
    ).fetchone()


def generate_monthly_report(
    db: sqlite3.Connection,
    start_date: str,
    end_date: str,
    daily_wage: float,
    include_photos: bool = True,
) -> io.BytesIO:
    wb = openpyxl.Workbook()

    # --- SHEET 1: DAILY ATTENDANCE (With Payment & Visual Evidence) ---
    ws_daily = wb.active
    ws_daily.title = "Daily Logs"

    headers = ["Date", "Worker Name", "Morning", "Evening", "Status", "Wages Earned"]
    if include_photos:
        headers += ["Morning Evidence", "Evening Evidence"]

    ws_daily.append(headers)

    # Style Headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="333333", end_color="333333", fill_type="solid")
    for col_num, _ in enumerate(headers, 1):
        cell = ws_daily.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Base URL so evidence cells can hyperlink back to the actual group photo
    base_url = None
    if include_photos:
        cfg = load_config()
        local_ip = get_local_ip()
        base_url = f"http://{local_ip}:{cfg['port']}"

    # Fetch Attendance with Joined Worker Names
    rows = db.execute(
        """
        SELECT a.*, w.name 
        FROM attendance a 
        JOIN workers w ON a.worker_id = w.id 
        WHERE a.date BETWEEN ? AND ?
        ORDER BY a.date DESC, w.name ASC
    """,
        (start_date, end_date),
    ).fetchall()

    row_idx = 2
    for r in rows:
        am = bool(r["morning_present"])
        pm = bool(r["evening_present"])

        # Payment Integrity Calculation
        if am and pm:
            status = "Full Day"
            wage_earned = daily_wage
        elif am or pm:
            status = "Half Day"
            wage_earned = daily_wage * 0.5
        else:
            status = "Absent"
            wage_earned = 0.0

        ws_daily.cell(row=row_idx, column=1, value=r["date"])
        ws_daily.cell(row=row_idx, column=2, value=r["name"])
        morning_cell = ws_daily.cell(row=row_idx, column=3, value="Present" if am else "-")
        evening_cell = ws_daily.cell(row=row_idx, column=4, value="Present" if pm else "-")
        ws_daily.cell(row=row_idx, column=5, value=status)
        ws_daily.cell(row=row_idx, column=6, value=wage_earned)

        # Embed Photo Evidence Directly into the Cell — separately per session
        if include_photos and (am or pm):
            ws_daily.row_dimensions[row_idx].height = 65

            for session, present, img_col, link_cell in (
                ("morning", am, "G", morning_cell),
                ("evening", pm, "H", evening_cell),
            ):
                if not present:
                    continue

                face_row = _find_face_evidence_for_session(db, r["worker_id"], r["date"], session)
                if not face_row:
                    continue

                # Hyperlink the Present cell back to the full original group photo
                photo_url = f"{base_url}/photos/{face_row['date']}/{face_row['session']}/{face_row['filename']}"
                link_cell.hyperlink = photo_url
                link_cell.font = Font(color="0563C1", underline="single")

                img_path = (
                    PHOTOS_DIR
                    / face_row["date"]
                    / face_row["session"]
                    / face_row["filename"]
                )
                img_bytes = _draw_face_circle(img_path, face_row["bbox"])
                if img_bytes:
                    xl_img = OpenPyxlImage(img_bytes)
                    ws_daily.add_image(xl_img, f"{img_col}{row_idx}")

        row_idx += 1

    # Adjust widths for readability
    ws_daily.column_dimensions["B"].width = 25
    ws_daily.column_dimensions["E"].width = 15
    ws_daily.column_dimensions["F"].width = 15
    if include_photos:
        ws_daily.column_dimensions["G"].width = 14
        ws_daily.column_dimensions["H"].width = 14

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def generate_report(
    start_date: str, end_date: str, daily_wage: float | None = None, include_photos: bool = True
) -> Path:
    """Generates an attendance Excel report and saves it to the REPORTS_DIR directory."""
    REPORTS_DIR.mkdir(parents=True, exist_ok=True)

    with get_connection() as conn:
        if daily_wage is None or daily_wage <= 0:
            repo = SettingsRepository(conn)
            daily_wage = float(repo.get("daily_wage", "500.0"))

        excel_bytes = generate_monthly_report(conn, start_date, end_date, daily_wage, include_photos=include_photos)

    filename = f"attendance_report_{start_date}_to_{end_date}.xlsx"
    report_path = REPORTS_DIR / filename

    with open(report_path, "wb") as f:
        f.write(excel_bytes.getbuffer())

    return report_path